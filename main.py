import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import matplotlib
import matplotlib.pyplot as plt
matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import os
import shutil

plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 数据模型定义
class Expense:
    def __init__(self, date, amount, category, description):
        self.date = date
        self.amount = amount
        self.category = category
        self.description = description

class Budget:
    def __init__(self, month, budget_amount):
        self.month = month
        self.budget_amount = budget_amount
        self.spent = 0.0
        self.remaining = budget_amount
    
    def update_spent(self, amount):
        self.spent += amount
        self.remaining = self.budget_amount - self.spent

# Excel文件操作类
class ExcelHandler:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = None
        self.ensure_worksheets_exist()
    
    def ensure_worksheets_exist(self):
        try:
            self.workbook = openpyxl.load_workbook(self.filename)
        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()
            self.workbook.remove(self.workbook.active)
            self.create_new_worksheets()
            self.save()
    
    def create_new_worksheets(self):
        # 创建消费记录工作表
        expense_sheet = self.workbook.create_sheet("消费记录")
        expense_sheet.append(["日期", "金额", "类别", "说明"])
        
        # 创建预算设置工作表
        budget_sheet = self.workbook.create_sheet("预算设置")
        budget_sheet.append(["月份", "预算金额", "已用金额", "剩余金额"])
        
        # 创建分类统计工作表
        category_sheet = self.workbook.create_sheet("分类统计")
        category_sheet.append(["类别", "总消费", "占比"])
    
    def save(self):
        self.workbook.save(self.filename)
    
    def add_expense(self, expense):
        expense_sheet = self.workbook["消费记录"]
        expense_sheet.append([
            expense.date.strftime("%Y-%m-%d"),
            expense.amount,
            expense.category,
            expense.description
        ])
        self.save()
    
    def get_all_expenses(self):
        expense_sheet = self.workbook["消费记录"]
        expenses = []
        for row in expense_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:  # 确保行不为空
                try:
                    date = datetime.strptime(row[0], "%Y-%m-%d")
                    amount = row[1]
                    category = row[2]
                    description = row[3]
                    expenses.append(Expense(date, amount, category, description))
                except:
                    continue
        return expenses
    
    def set_budget(self, month, budget_amount):
        budget_sheet = self.workbook["预算设置"]
        # 检查是否已有该月份的预算记录
        for row_idx, row in enumerate(budget_sheet.iter_rows(min_row=2, values_only=True), 2):
            if row[0] == month:
                # 更新现有预算
                budget_sheet.cell(row=row_idx, column=2, value=budget_amount)
                budget_sheet.cell(row=row_idx, column=3, value=0.0)
                budget_sheet.cell(row=row_idx, column=4, value=budget_amount)
                self.save()
                return
        # 如果没有，则添加新预算记录
        budget_sheet.append([month, budget_amount, 0.0, budget_amount])
        self.save()
    
    def get_budget(self, month):
        budget_sheet = self.workbook["预算设置"]
        for row in budget_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == month:
                budget = Budget(month, row[1])
                budget.spent = row[2]
                budget.remaining = row[3]
                return budget
        return None
    
    def update_budget_spent(self, month, amount):
        budget_sheet = self.workbook["预算设置"]
        for row_idx, row in enumerate(budget_sheet.iter_rows(min_row=2, values_only=True), 2):
            if row[0] == month:
                new_spent = row[2] + amount
                new_remaining = row[1] - new_spent
                budget_sheet.cell(row=row_idx, column=3, value=new_spent)
                budget_sheet.cell(row=row_idx, column=4, value=new_remaining)
                self.save()
                return
    
    def generate_category_statistics(self):
        expense_sheet = self.workbook["消费记录"]
        category_sheet = self.workbook["分类统计"]
        
        # 清空现有统计数据
        if category_sheet.max_row > 1:
            category_sheet.delete_rows(2, category_sheet.max_row - 1)
        
        # 计算各类别总消费
        categories = {}
        for row in expense_sheet.iter_rows(min_row=2, values_only=True):
            if row[1] is not None and row[2] is not None:  # 确保金额和类别不为空
                category = row[2]
                amount = row[1]
                if category in categories:
                    categories[category] += amount
                else:
                    categories[category] = amount
        
        # 计算总消费和占比
        total_spent = sum(categories.values())
        for category, total in categories.items():
            percentage = (total / total_spent) * 100 if total_spent != 0 else 0
            category_sheet.append([category, total, f"{percentage:.1f}%"])
        
        self.save()
    
    def get_category_statistics(self):
        category_sheet = self.workbook["分类统计"]
        statistics = []
        for row in category_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:  # 确保类别不为空
                category = row[0]
                total = row[1]
                percentage = row[2]
                statistics.append((category, total, percentage))
        return statistics

# 添加消费记录窗口
class AddExpenseWindow:
    def __init__(self, master, excel_handler, main_gui):
        self.master = master
        self.master.title("添加消费记录")
        self.excel_handler = excel_handler
        self.main_gui = main_gui
        
        # 设置窗口大小和位置
        window_width = 300
        window_height = 200
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()
        x_position = (screen_width - window_width) // 2
        y_position = (screen_height - window_height) // 2
        self.master.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")
        
        # 创建输入组件
        self.create_widgets()
    
    def create_widgets(self):
        # 金额输入
        ttk.Label(self.master, text="金额:").grid(row=0, column=0, padx=5, pady=5)
        self.amount_entry = ttk.Entry(self.master)
        self.amount_entry.grid(row=0, column=1, padx=5, pady=5)
        
        # 类别选择
        ttk.Label(self.master, text="类别:").grid(row=1, column=0, padx=5, pady=5)
        self.category_combobox = ttk.Combobox(self.master, values=[
            "餐饮", "交通", "学习", "零食", "日用品", "娱乐",
        ], state="readonly")
        self.category_combobox.set("餐饮")
        self.category_combobox.grid(row=1, column=1, padx=5, pady=5)
        
        # 说明输入
        ttk.Label(self.master, text="说明:").grid(row=2, column=0, padx=5, pady=5)
        self.description_entry = ttk.Entry(self.master)
        self.description_entry.grid(row=2, column=1, padx=5, pady=5)
        
        # 确定按钮
        ttk.Button(self.master, text="确定", command=self.save_expense).grid(
            row=3,
            column=0,
            columnspan=2,
            pady=5
        )
    
    def save_expense(self):
        try:
            amount = float(self.amount_entry.get())
            category = self.category_combobox.get()
            description = self.description_entry.get()
            date = datetime.now()
            
            if amount <= 0:
                raise ValueError("金额必须大于0！")
            
            expense = Expense(date, amount, category, description)
            self.main_gui.add_new_expense(expense)
            self.master.destroy()
        except ValueError as e:
            messagebox.showerror("错误", str(e))

# 恢复备份窗口
class RestoreBackupWindow:
    def __init__(self, master, backup_dir, main_gui):
        self.master = master
        self.master.title("恢复备份")
        self.backup_dir = backup_dir
        self.main_gui = main_gui
        
        # 设置窗口大小和位置
        window_width = 400
        window_height = 300
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()
        x_position = (screen_width - window_width) // 2
        y_position = (screen_height - window_height) // 2
        self.master.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")
        
        # 创建列表框显示备份文件
        self.backup_listbox = tk.Listbox(self.master, width=50)
        self.backup_listbox.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(self.backup_listbox, orient=tk.VERTICAL, command=self.backup_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.backup_listbox.config(yscrollcommand=scrollbar.set)
        
        # 加载备份文件列表
        self.load_backup_files()
        
        # 恢复按钮
        ttk.Button(self.master, text="恢复选中的备份", command=self.restore_selected).pack(pady=10)
    
    def load_backup_files(self):
        backup_files = sorted(os.listdir(self.backup_dir), reverse=True)  # 按时间倒序排列
        for file in backup_files:
            if file.endswith('.xlsx'):
                self.backup_listbox.insert(tk.END, file)
    
    def restore_selected(self):
        selection = self.backup_listbox.curselection()
        if not selection:
            messagebox.showerror("错误", "请选择一个备份文件！")
            return
        
        selected_file = self.backup_listbox.get(selection[0])
        backup_path = os.path.join(self.backup_dir, selected_file)
        
        try:
            # 先关闭主程序的Excel文件
            self.main_gui.excel_handler.workbook.close()
            
            # 复制备份文件到主文件
            shutil.copyfile(backup_path, "expenses.xlsx")
            
            # 重新打开Excel文件
            self.main_gui.excel_handler = ExcelHandler("expenses.xlsx")
            
            # 刷新数据
            self.main_gui.refresh_data()
            
            messagebox.showinfo("恢复成功", "数据已从备份恢复！")
            self.master.destroy()
        except Exception as e:
            messagebox.showerror("恢复失败", f"恢复过程中出错：{str(e)}")
            # 重新打开Excel文件
            self.main_gui.excel_handler = ExcelHandler("expenses.xlsx")

# 主窗口界面
class ExpenseManagerGUI:
    def __init__(self, master):
        self.master = master
        self.master.title("大学生消费管理系统")
        
        # 设置应用主题
        ttk.Style().theme_use('clam')
        
        # 初始化Excel处理器
        self.excel_handler = ExcelHandler("expenses.xlsx")
        
        # 设置窗口大小和位置
        window_width = 800
        window_height = 600
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()
        x_position = (screen_width - window_width) // 2
        y_position = (screen_height - window_height) // 2
        self.master.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")
        
        # 创建主框架
        self.main_frame = ttk.Frame(self.master, padding="10")
        self.main_frame.grid(row=0, column=0, sticky=(tk.N, tk.W, tk.E, tk.S))
        self.master.columnconfigure(0, weight=1)
        self.master.rowconfigure(0, weight=1)
        
        # 创建界面组件
        self.create_widgets()
        
        # 初始化预算提醒定时器
        self.check_budget_alert()
        
        # 加载消费记录
        self.load_expenses()
    
    def create_widgets(self):
        # 创建顶部预算信息区域
        self.create_budget_section()
        
        # 创建搜索区域
        self.create_search_section()
        
        # 创建左侧消费记录区域
        self.create_expense_list_section()
        
        # 创建右侧图表区域
        self.create_chart_section()
        
        # 创建底部操作按钮
        self.create_action_buttons()
        
        # 创建备份与恢复区域
        self.create_backup_section()
        
        # 创建预算提醒设置区域
        self.create_budget_alert_settings()
        
        # 添加工具提示
        self.create_tooltips()
    
    def create_budget_section(self):
        # 预算设置框架
        budget_frame = ttk.LabelFrame(self.main_frame, text="月度预算设置")
        budget_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=5, pady=5)
        
        # 月份选择
        self.month_combobox = ttk.Combobox(budget_frame, values=[
            "一月", "二月", "三月", "四月", "五月", "六月",
            "七月", "八月", "九月", "十月", "十一月", "十二月"
        ], state="readonly")
        current_month_number = datetime.now().month
        month_names = ["一月", "二月", "三月", "四月", "五月", "六月",
                      "七月", "八月", "九月", "十月", "十一月", "十二月"]
        self.month_combobox.set(month_names[current_month_number - 1])  # 默认显示当前月份
        self.month_combobox.grid(row=0, column=0, padx=5, pady=5)
        
        # 预算金额输入
        self.budget_entry = ttk.Entry(budget_frame)
        self.budget_entry.grid(row=0, column=1, padx=5, pady=5)
        
        # 设置预算按钮
        self.set_budget_button = ttk.Button(budget_frame, text="设置预算", command=self.set_budget)
        self.set_budget_button.grid(row=0, column=2, padx=5, pady=5)
        
        # 预算状态显示
        self.budget_status_label = ttk.Label(budget_frame, text="")
        self.budget_status_label.grid(row=1, column=0, columnspan=3, sticky=tk.W, padx=5, pady=5)
        
        # 加载当前月份预算
        current_month = month_names[current_month_number - 1]
        self.load_budget(current_month)
    
    def create_search_section(self):
        search_frame = ttk.LabelFrame(self.main_frame, text="消费记录查询")
        search_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=5, pady=5)
        
        self.search_entry = ttk.Entry(search_frame)
        self.search_entry.grid(row=0, column=0, padx=5, pady=5, sticky=tk.EW)
        
        self.search_button = ttk.Button(search_frame, text="搜索", command=self.search_expenses)
        self.search_button.grid(row=0, column=1, padx=5, pady=5)
        
        self.search_type = ttk.Combobox(search_frame, values=["全部", "日期", "类别", "说明"], state="readonly", width=10)
        self.search_type.set("全部")
        self.search_type.grid(row=0, column=2, padx=5, pady=5)
        
        # 设置列权重，使搜索框可扩展
        search_frame.columnconfigure(0, weight=1)
    
    def create_expense_list_section(self):
        # 消费记录框架
        expense_list_frame = ttk.LabelFrame(self.main_frame, text="消费记录")
        expense_list_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=5, pady=5)
        
        # 消费记录表格
        self.expense_tree = ttk.Treeview(expense_list_frame, columns=("金额", "类别", "说明", "日期"), show="headings")
        self.expense_tree.heading("金额", text="金额")
        self.expense_tree.heading("类别", text="类别")
        self.expense_tree.heading("说明", text="说明")
        self.expense_tree.heading("日期", text="日期")
        
        # 设置列宽
        self.expense_tree.column("金额", width=80, anchor=tk.CENTER)
        self.expense_tree.column("类别", width=80, anchor=tk.CENTER)
        self.expense_tree.column("说明", width=150, anchor=tk.W)
        self.expense_tree.column("日期", width=100, anchor=tk.CENTER)
        
        self.expense_tree.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E), padx=5, pady=5)
        
        # 添加垂直滚动条
        scrollbar = ttk.Scrollbar(expense_list_frame, orient=tk.VERTICAL, command=self.expense_tree.yview)
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.expense_tree.configure(yscrollcommand=scrollbar.set)
        
        # 使表格随窗口大小调整
        expense_list_frame.grid_rowconfigure(0, weight=1)
        expense_list_frame.grid_columnconfigure(0, weight=1)
    
    def create_chart_section(self):
        # 图表框架
        chart_frame = ttk.LabelFrame(self.main_frame, text="消费分类统计")
        chart_frame.grid(row=0, column=1, rowspan=4, sticky=(tk.N, tk.S, tk.W, tk.E), padx=5, pady=5)
        
        # 预留图表显示区域
        self.chart_canvas = tk.Canvas(chart_frame, width=400, height=300)
        self.chart_canvas.grid(row=0, column=0, padx=5, pady=5)
        
        # 使图表区域可调整大小
        chart_frame.grid_rowconfigure(0, weight=1)
        chart_frame.grid_columnconfigure(0, weight=1)
        
        # 初始化图表
        self.update_chart()
    
    def create_action_buttons(self):
        # 操作按钮框架
        button_frame = ttk.Frame(self.main_frame)
        button_frame.grid(row=3, column=0, sticky=tk.W, padx=5, pady=5)
        
        # 添加消费按钮
        self.add_expense_button = ttk.Button(button_frame, text="添加消费", command=self.open_add_expense_window)
        self.add_expense_button.grid(row=0, column=0, padx=5, pady=5)
        
        # 生成月度报告按钮
        self.generate_report_button = ttk.Button(button_frame, text="生成月度报告", command=self.generate_monthly_report)
        self.generate_report_button.grid(row=0, column=1, padx=5, pady=5)
        
        # 刷新数据按钮
        self.refresh_button = ttk.Button(button_frame, text="刷新数据", command=self.refresh_data)
        self.refresh_button.grid(row=0, column=2, padx=5, pady=5)
        
        # 退出系统按钮
        self.quit_button = ttk.Button(button_frame, text="退出系统", command=self.master.destroy)
        self.quit_button.grid(row=0, column=3, padx=5, pady=5)
    
    def create_backup_section(self):
        backup_frame = ttk.LabelFrame(self.main_frame, text="数据备份与恢复")
        backup_frame.grid(row=4, column=0, sticky=tk.W, padx=5, pady=5)
        
        ttk.Button(backup_frame, text="创建备份", command=self.create_backup).grid(
            row=0,
            column=0,
            padx=5,
            pady=5
        )
        
        ttk.Button(backup_frame, text="恢复备份", command=self.restore_backup).grid(
            row=0,
            column=1,
            padx=5,
            pady=5
        )
    
    def create_budget_alert_settings(self):
        alert_frame = ttk.LabelFrame(self.main_frame, text="预算提醒设置")
        alert_frame.grid(row=4, column=1, sticky=tk.W, padx=5, pady=5)
        
        ttk.Label(alert_frame, text="提醒阈值:").grid(row=0, column=0, padx=5, pady=5)
        self.alert_threshold = ttk.Entry(alert_frame, width=5)
        self.alert_threshold.grid(row=0, column=1, padx=5, pady=5)
        self.alert_threshold.insert(0, "90")  # 默认90%
        
        ttk.Label(alert_frame, text="%").grid(row=0, column=2, padx=0, pady=5, sticky=tk.W)
        
        ttk.Button(alert_frame, text="保存设置", command=self.save_alert_settings).grid(
            row=0,
            column=3,
            padx=5,
            pady=5
        )
        
        # 加载现有设置（如果有）
        self.load_alert_settings()
    
    def create_tooltips(self):
        # 这里使用简化的工具提示实现，不使用messagebox以避免频繁弹窗
        pass
    
    def load_budget(self, month):
        budget = self.excel_handler.get_budget(month)
        if budget:
            self.budget_entry.delete(0, tk.END)
            self.budget_entry.insert(0, budget.budget_amount)
            self.budget_status_label.config(text=f"已用: {budget.spent:.1f} 元, 剩余: {budget.remaining:.1f} 元")
        else:
            self.budget_entry.delete(0, tk.END)
            self.budget_status_label.config(text="")
    
    def set_budget(self):
        month = self.month_combobox.get()
        try:
            budget_amount = float(self.budget_entry.get())
            if budget_amount <= 0:
                raise ValueError("预算金额必须大于0！")
            self.excel_handler.set_budget(month, budget_amount)
            self.load_budget(month)
            messagebox.showinfo("预算设置", "预算设置成功！")
        except ValueError as e:
            messagebox.showerror("错误", str(e))
    
    def open_add_expense_window(self):
        add_window = tk.Toplevel(self.master)
        AddExpenseWindow(add_window, self.excel_handler, self)
    
    def add_new_expense(self, expense):
        self.excel_handler.add_expense(expense)
        month_names = ["一月", "二月", "三月", "四月", "五月", "六月",
                      "七月", "八月", "九月", "十月", "十一月", "十二月"]
        current_month = month_names[expense.date.month - 1]
        self.excel_handler.update_budget_spent(current_month, expense.amount)
        self.load_expenses()
        self.update_chart()
        self.check_budget_alert()
    
    def load_expenses(self):
        # 清空现有数据
        for item in self.expense_tree.get_children():
            self.expense_tree.delete(item)
        
        # 加载新数据
        expenses = self.excel_handler.get_all_expenses()
        for expense in expenses:
            self.expense_tree.insert(
                "",
                tk.END,
                values=(
                    f"{expense.amount:.1f} 元",
                    expense.category,
                    expense.description,
                    expense.date.strftime("%Y-%m-%d")
                )
            )
    
    def search_expenses(self):
        query = self.search_entry.get().strip().lower()
        search_type = self.search_type.get()
        
        # 清空现有数据
        for item in self.expense_tree.get_children():
            self.expense_tree.delete(item)
        
        # 加载匹配数据
        expenses = self.excel_handler.get_all_expenses()
        for expense in expenses:
            match = False
            if search_type == "全部":
                if query in str(expense.amount).lower() or \
                   query in expense.category.lower() or \
                   query in expense.description.lower() or \
                   query in expense.date.strftime("%Y-%m-%d").lower():
                    match = True
            elif search_type == "日期":
                if query in expense.date.strftime("%Y-%m-%d").lower():
                    match = True
            elif search_type == "类别":
                if query in expense.category.lower():
                    match = True
            elif search_type == "说明":
                if query in expense.description.lower():
                    match = True
            
            if match:
                self.expense_tree.insert(
                    "",
                    tk.END,
                    values=(
                        f"{expense.amount:.1f} 元",
                        expense.category,
                        expense.description,
                        expense.date.strftime("%Y-%m-%d")
                    )
                )
    
    def update_chart(self):
        # 生成消费分类统计
        self.excel_handler.generate_category_statistics()
        statistics = self.excel_handler.get_category_statistics()
        
        # 准备图表数据
        categories = [stat[0] for stat in statistics]
        amounts = [stat[1] for stat in statistics]
        
        # 清除现有图表
        self.chart_canvas.delete("all")
        
        # 创建新图表
        if categories:
            fig = Figure(figsize=(3, 2), dpi=100)
            ax = fig.add_subplot(111)
            ax.pie(amounts, labels=categories, autopct='%1.1f%%', startangle=90)
            ax.axis('equal')  # 保证饼图为正圆形
            
            canvas = FigureCanvasTkAgg(fig, master=self.chart_canvas)
            canvas.draw()
            canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
            
            # 将图表嵌入到Tkinter窗口中
            self.chart_canvas.create_window((0,0), window=canvas.get_tk_widget(), anchor=tk.NW)
        else:
            self.chart_canvas.create_text(
                200, 150,
                text="无消费记录，无法生成图表",
                fill="gray",
                font=("Arial", 12)
            )
    
    def check_budget_alert(self):
        try:
            with open("alert_settings.txt", "r") as f:
                threshold = int(f.read().strip())
        except (FileNotFoundError, ValueError):
            threshold = 90  # 默认值
        
        month_names = ["一月", "二月", "三月", "四月", "五月", "六月",
                      "七月", "八月", "九月", "十月", "十一月", "十二月"]
        current_month = month_names[datetime.now().month - 1]
        budget = self.excel_handler.get_budget(current_month)
        
        if budget:
            threshold_amount = budget.budget_amount * threshold / 100
            if budget.spent >= threshold_amount and not hasattr(self, 'alert_shown'):
                messagebox.showwarning("预算提醒", f"已接近月度预算限额（已用{budget.spent:.1f}元，剩余{budget.remaining:.1f}元）！")
                self.alert_shown = True
            elif budget.spent < threshold_amount:
                self.alert_shown = False  # 重置提醒状态
        
        # 每小时检查一次预算状态
        self.master.after(3600000, self.check_budget_alert)
    
    def refresh_data(self):
        self.load_expenses()
        self.update_chart()
        self.load_budget(self.month_combobox.get())
    
    def generate_monthly_report(self):
        from report_generator import generate_monthly_report
        month_names = ["一月", "二月", "三月", "四月", "五月", "六月",
                      "七月", "八月", "九月", "十月", "十一月", "十二月"]
        current_month = month_names[datetime.now().month - 1]
        
        try:
            generate_monthly_report(current_month, self.excel_handler)
            messagebox.showinfo("报告生成", "月度报告已生成！")
        except Exception as e:
            messagebox.showerror("报告生成失败", f"生成报告时出错：{str(e)}")
    
    def create_backup(self):
        backup_dir = "expenses_backup"
        os.makedirs(backup_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        backup_filename = f"expenses_{timestamp}.xlsx"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        try:
            # 关闭Excel文件后再备份
            self.excel_handler.workbook.close()
            shutil.copyfile("expenses.xlsx", backup_path)
            # 重新打开Excel文件
            self.excel_handler = ExcelHandler("expenses.xlsx")
            messagebox.showinfo("备份成功", f"备份文件已保存至：{backup_path}")
        except Exception as e:
            messagebox.showerror("备份失败", f"备份过程中出错：{str(e)}")
            # 重新打开Excel文件
            self.excel_handler = ExcelHandler("expenses.xlsx")
    
    def restore_backup(self):
        backup_dir = "expenses_backup"
        if not os.path.exists(backup_dir) or not os.listdir(backup_dir):
            messagebox.showerror("无备份", "没有可用的备份文件！")
            return
        
        # 创建选择备份文件的窗口
        restore_window = tk.Toplevel(self.master)
        RestoreBackupWindow(restore_window, backup_dir, self)
    
    def load_alert_settings(self):
        try:
            with open("alert_settings.txt", "r") as f:
                threshold = f.read().strip()
                if threshold:
                    self.alert_threshold.delete(0, tk.END)
                    self.alert_threshold.insert(0, threshold)
        except FileNotFoundError:
            pass
    
    def save_alert_settings(self):
        try:
            threshold = int(self.alert_threshold.get())
            if 0 <= threshold <= 100:
                with open("alert_settings.txt", "w") as f:
                    f.write(str(threshold))
                messagebox.showinfo("设置保存", "提醒阈值设置成功！")
            else:
                raise ValueError("阈值必须在0到100之间！")
        except ValueError as e:
            messagebox.showerror("错误", str(e))

# 主程序入口
if __name__ == "__main__":
    root = tk.Tk()
    app = ExpenseManagerGUI(root)
    root.mainloop()